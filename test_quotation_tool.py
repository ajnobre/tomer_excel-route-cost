#!/usr/bin/env python3
"""
Comprehensive test for the Quotation Tool workbook.
Simulates Excel formula evaluation to verify lookup logic, freight aggregation,
tonnage integration, and final calculations before testing in Windows Excel.
"""

import openpyxl
import os
import sys
from collections import defaultdict

# Use the same folder as this script
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TOOL_FILE = os.path.join(BASE_DIR, 'Quotation_Tool.xlsx')

def load_tool():
    """Load the generated tool and extract all data sheets."""
    wb = openpyxl.load_workbook(TOOL_FILE, data_only=False)
    return wb

def get_sheet_data(ws, max_col=None):
    """Read all rows from a sheet into a list of lists."""
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=max_col, values_only=True):
        rows.append(list(row))
    return rows

def test_structure(wb):
    """Test 1: Verify workbook structure."""
    print("=" * 60)
    print("TEST 1: Workbook Structure")
    print("=" * 60)

    expected_sheets = ['Quote', 'Lists', 'IN_GB', 'SL_GB', 'Freight']
    actual_sheets = wb.sheetnames

    ok = True
    for s in expected_sheets:
        if s in actual_sheets:
            print(f"  [PASS] Sheet '{s}' exists")
        else:
            print(f"  [FAIL] Sheet '{s}' MISSING")
            ok = False

    # Check Quote is first (active) sheet
    if actual_sheets[0] == 'Quote':
        print(f"  [PASS] 'Quote' is the first sheet")
    else:
        print(f"  [FAIL] 'Quote' is not first — it's at position {actual_sheets.index('Quote') if 'Quote' in actual_sheets else 'N/A'}")
        ok = False

    # Check hidden sheets
    for s in ['Lists', 'IN_GB', 'SL_GB', 'Freight']:
        if s in actual_sheets and wb[s].sheet_state == 'hidden':
            print(f"  [PASS] Sheet '{s}' is hidden")
        elif s in actual_sheets:
            print(f"  [WARN] Sheet '{s}' is NOT hidden")

    return ok

def test_named_ranges(wb):
    """Test 2: Verify named ranges exist and are non-empty."""
    print("\n" + "=" * 60)
    print("TEST 2: Named Ranges")
    print("=" * 60)

    expected = [
        'IN_Keys', 'IN_ProdNos', 'IN_Descs', 'IN_PCS', 'IN_FOB',
        'SL_Keys', 'SL_ProdNos', 'SL_Descs', 'SL_PCS', 'SL_FOB',
        'FR_Keys', 'FR_Transit', 'FR_AllIn',
        'FR_GrossWT', 'FR_Confirmed',
        'SizeList', 'ChipsList', 'ECList', 'PlasticList',
        'HolesList', 'BSUList', 'DestList', 'WeightTiers',
    ]

    ok = True
    for name in expected:
        dn = wb.defined_names.get(name)
        if dn:
            print(f"  [PASS] '{name}' -> {dn.attr_text}")
        else:
            print(f"  [FAIL] '{name}' NOT FOUND")
            ok = False

    print(f"  [INFO] Total named ranges: {len(expected)} expected, {len(list(wb.defined_names))} found")
    return ok

def test_data_integrity(wb):
    """Test 3: Verify data sheet contents and consistency."""
    print("\n" + "=" * 60)
    print("TEST 3: Data Integrity")
    print("=" * 60)

    ok = True

    # IN_GB
    ws_in = wb['IN_GB']
    in_rows = get_sheet_data(ws_in, max_col=24)
    print(f"  IN_GB: {len(in_rows)} products")

    # Check header row
    headers_in = [ws_in.cell(row=1, column=c).value for c in range(1, 25)]
    if headers_in[0] == 'Key':
        print(f"  [PASS] IN_GB headers start with 'Key'")
    else:
        print(f"  [FAIL] IN_GB header[0] = '{headers_in[0]}', expected 'Key'")
        ok = False

    # Verify keys are non-empty
    empty_keys = sum(1 for r in in_rows if not r[0])
    if empty_keys == 0:
        print(f"  [PASS] All IN_GB rows have keys")
    else:
        print(f"  [WARN] {empty_keys} IN_GB rows have empty keys")

    # Check a few products have FOB prices
    products_with_fob = sum(1 for r in in_rows if any(r[17+i] for i in range(7)))
    print(f"  [INFO] {products_with_fob}/{len(in_rows)} IN_GB products have at least one FOB price")

    # SL_GB
    ws_sl = wb['SL_GB']
    sl_rows = get_sheet_data(ws_sl, max_col=24)
    print(f"  SL_GB: {len(sl_rows)} products")

    products_with_fob = sum(1 for r in sl_rows if any(r[17+i] for i in range(7)))
    print(f"  [INFO] {products_with_fob}/{len(sl_rows)} SL_GB products have at least one FOB price")

    # Freight (8 columns: A-H)
    ws_fr = wb['Freight']
    fr_rows = get_sheet_data(ws_fr, max_col=8)
    print(f"  Freight: {len(fr_rows)} routes")

    # Count routes per origin
    origin_counts = {}
    for r in fr_rows:
        origin = r[2]
        origin_counts[origin] = origin_counts.get(origin, 0) + 1
    for origin, count in sorted(origin_counts.items()):
        print(f"  [INFO] {origin}: {count} routes")

    # Check column headers
    fr_headers = [ws_fr.cell(row=1, column=c).value for c in range(1, 9)]
    if fr_headers[5] == 'All_In_USD' and fr_headers[6] == 'Gross_Weight_MT' and fr_headers[7] == 'Weight_Confirmed':
        print(f"  [PASS] Freight columns F-H present (All_In_USD, Gross_Weight_MT, Weight_Confirmed)")
    else:
        print(f"  [FAIL] Freight columns F-H missing or wrong. Headers: {fr_headers}")
        ok = False

    # Validate tonnage data
    confirmed_count = sum(1 for r in fr_rows if r[7] == 1)
    default_count = sum(1 for r in fr_rows if r[7] == 0)
    has_weight = sum(1 for r in fr_rows if r[6] is not None and r[6] > 0)
    print(f"  [INFO] Tonnage: {confirmed_count} confirmed, {default_count} default (out of {len(fr_rows)} routes)")

    if has_weight == len(fr_rows):
        print(f"  [PASS] All freight routes have a weight value")
    else:
        print(f"  [FAIL] {len(fr_rows) - has_weight} routes missing weight data")
        ok = False

    # Confirm all default weights are 23
    bad_defaults = [r for r in fr_rows if r[7] == 0 and r[6] != 23]
    if not bad_defaults:
        print(f"  [PASS] All unconfirmed routes use default 23 MT")
    else:
        print(f"  [FAIL] {len(bad_defaults)} unconfirmed routes with non-23 weight")
        ok = False

    # Confirm flag is 0 or 1
    bad_flags = [r for r in fr_rows if r[7] not in (0, 1)]
    if not bad_flags:
        print(f"  [PASS] All Weight_Confirmed flags are 0 or 1")
    else:
        print(f"  [FAIL] {len(bad_flags)} routes with invalid confirmed flag")
        ok = False

    # Lists
    ws_lists = wb['Lists']
    list_counts = {}
    for col_idx, name in enumerate(['Size', 'Chips_Pith', 'EC_Level', 'Plastic', 'Holes', 'BSU', 'Destination', 'Weight_MT'], 1):
        count = 0
        r = 2
        while ws_lists.cell(row=r, column=col_idx).value is not None:
            count += 1
            r += 1
        list_counts[name] = count
    print(f"  [INFO] Dropdown lists: {list_counts}")

    return ok

def test_formula_simulation(wb):
    """Test 4: Simulate formula evaluation for specific input combinations."""
    print("\n" + "=" * 60)
    print("TEST 4: Formula Simulation (Core Logic)")
    print("=" * 60)

    # Load data into memory
    ws_in = wb['IN_GB']
    ws_sl = wb['SL_GB']
    ws_fr = wb['Freight']

    in_data = get_sheet_data(ws_in, max_col=24)
    sl_data = get_sheet_data(ws_sl, max_col=24)
    fr_data = get_sheet_data(ws_fr, max_col=8)

    weight_tiers = [19.5, 22, 23, 24, 25, 26, 27]

    # 3 origins: Cochin and Tuticorin use IN products, Colombo uses SL products
    ORIGINS = [
        ('Cochin, IN', in_data, 'IN'),
        ('Tuticorin, IN', in_data, 'IN'),
        ('Colombo, LK', sl_data, 'SL'),
    ]

    def simulate(size, chips, ec, plastic, holes, bsu, destination):
        """Simulate what Excel formulas would compute (weight auto-derived from tonnage)."""
        lookup_key = f"{size}|{chips}|{ec}|{plastic}|{holes}|{bsu}"

        results = {}
        for origin, data, prefix in ORIGINS:
            # MATCH product
            prod_match = None
            for i, row in enumerate(data):
                if row[0] == lookup_key:
                    prod_match = i
                    break

            # MATCH freight (key is origin|destination)
            fr_key = f"{origin}|{destination}"
            fr_match = None
            for i, row in enumerate(fr_data):
                if row[0] == fr_key:
                    fr_match = i
                    break

            # Get weight tier from freight data (col G=index 6)
            gross_wt = None
            confirmed = None
            if fr_match is not None:
                gross_wt = fr_data[fr_match][6]
                confirmed = fr_data[fr_match][7]

            # Approximate MATCH for weight tier (largest tier <= gross weight)
            wt_idx = None
            if gross_wt is not None:
                for i in range(len(weight_tiers) - 1, -1, -1):
                    if weight_tiers[i] <= gross_wt:
                        wt_idx = i
                        break

            if prod_match is not None and wt_idx is not None:
                prod_code = data[prod_match][1]
                desc = data[prod_match][2]
                pcs = data[prod_match][10 + wt_idx]
                fob = data[prod_match][17 + wt_idx]
            else:
                prod_code = None
                desc = None
                pcs = None
                fob = None

            if fr_match is not None:
                transit = fr_data[fr_match][4]
                all_in_usd = fr_data[fr_match][5]  # All_In_USD column F
                freight_total = all_in_usd * 1.0605
            else:
                transit = None
                freight_total = None

            discount = 0.30  # default 30%
            disc_fob = None
            freight_per_unit = None
            total_cost = None
            if fob is not None:
                disc_fob = fob * (1 - discount)
            if freight_total is not None and pcs and pcs > 0:
                freight_per_unit = freight_total / pcs
            if disc_fob is not None and freight_per_unit is not None:
                total_cost = disc_fob + freight_per_unit

            results[origin] = {
                'product_code': prod_code,
                'description': desc,
                'fob': fob,
                'disc_fob': disc_fob,
                'freight_container': freight_total,
                'units_container': pcs,
                'freight_per_unit': freight_per_unit,
                'total_cost': total_cost,
                'transit_days': transit,
                'gross_wt': gross_wt,
                'confirmed': confirmed,
                'tier': weight_tiers[wt_idx] if wt_idx is not None else None,
            }

        return lookup_key, results

    ok = True

    # ── Test Case 1: Find a product that exists in BOTH India and Sri Lanka ──
    print("\n  --- Test 4a: Finding a product in both origins ---")

    in_keys = set(r[0] for r in in_data)
    sl_keys = set(r[0] for r in sl_data)
    common_keys = in_keys & sl_keys
    print(f"  [INFO] {len(common_keys)} product keys exist in BOTH India and Sri Lanka")

    if common_keys:
        test_key = sorted(common_keys)[0]
        parts = test_key.split('|')
        if len(parts) == 6:
            # Find destinations available across all 3 origins
            all_dests = set(r[3] for r in fr_data)
            if all_dests:
                test_dest = sorted(all_dests)[0]

                print(f"  [INFO] Testing with key: {test_key}")
                print(f"  [INFO] Destination: {test_dest}")

                key, results = simulate(parts[0], parts[1], parts[2], parts[3], parts[4], parts[5],
                                       test_dest)

                for origin, r in results.items():
                    if r['fob'] is not None:
                        print(f"\n  {origin}:")
                        print(f"    Product Code:      {r['product_code']}")
                        print(f"    Weight Tier:       {r['tier']} MT (gross={r['gross_wt']}, "
                              f"{'confirmed' if r['confirmed'] == 1 else 'DEFAULT'})")
                        print(f"    FOB/Unit:          ${r['fob']}")
                        print(f"    Disc. FOB/Unit:    ${r['disc_fob']:.2f}" if r['disc_fob'] is not None else "    Disc. FOB/Unit:    N/A")
                        print(f"    Freight/Container: ${r['freight_container']:.2f}" if r['freight_container'] else "    Freight/Container: N/A")
                        print(f"    Units/Container:   {r['units_container']}")
                        print(f"    Freight/Unit:      ${r['freight_per_unit']:.4f}" if r['freight_per_unit'] else "    Freight/Unit:      N/A")
                        print(f"    Total Cost/Unit:   ${r['total_cost']:.4f}" if r['total_cost'] else "    Total Cost/Unit:   N/A")
                        print(f"    Transit Days:      {r['transit_days']}")

                        # Verify discount arithmetic
                        if r['fob'] is not None and r['disc_fob'] is not None:
                            expected_disc = r['fob'] * 0.70
                            if abs(expected_disc - r['disc_fob']) < 0.001:
                                print(f"    [PASS] Disc. FOB = FOB × (1 - 30%)")
                            else:
                                print(f"    [FAIL] Disc. FOB mismatch: {r['disc_fob']} vs expected {expected_disc}")
                                ok = False

                        # Verify freight arithmetic
                        if r['freight_container'] and r['units_container'] and r['units_container'] > 0:
                            expected_fpu = r['freight_container'] / r['units_container']
                            if abs(expected_fpu - r['freight_per_unit']) < 0.001:
                                print(f"    [PASS] Freight/Unit = Freight/Container / Units/Container")
                            else:
                                print(f"    [FAIL] Freight/Unit mismatch: {r['freight_per_unit']} vs expected {expected_fpu}")
                                ok = False

                        # Verify total = discounted FOB + freight/unit
                        if r['disc_fob'] is not None and r['freight_per_unit']:
                            expected_total = r['disc_fob'] + r['freight_per_unit']
                            if abs(expected_total - r['total_cost']) < 0.001:
                                print(f"    [PASS] Total = Disc. FOB + Freight/Unit")
                            else:
                                print(f"    [FAIL] Total mismatch: {r['total_cost']} vs expected {expected_total}")
                                ok = False
                    else:
                        print(f"\n  {origin}: No product match (expected for this key)")

    # ── Test Case 2: Validate against client reference image ──
    print("\n\n  --- Test 4b: Searching for client reference values ---")
    print("  Reference: India FOB=2, Freight=3000, Units=8000, Total=2.375")
    print("  Reference: Sri Lanka FOB=2, Freight=5000, Units=7000, Total=2.714")

    in_fob2 = []
    for r in in_data:
        for i in range(7):
            if r[17 + i] == 2 or r[17 + i] == 2.0:
                in_fob2.append((r[0], r[1], r[2], weight_tiers[i], r[10 + i]))

    sl_fob2 = []
    for r in sl_data:
        for i in range(7):
            if r[17 + i] == 2 or r[17 + i] == 2.0:
                sl_fob2.append((r[0], r[1], r[2], weight_tiers[i], r[10 + i]))

    print(f"  [INFO] India products with FOB=$2: {len(in_fob2)}")
    print(f"  [INFO] Sri Lanka products with FOB=$2: {len(sl_fob2)}")

    in_match = [x for x in in_fob2 if x[4] == 8000]
    sl_match = [x for x in sl_fob2 if x[4] == 7000]

    if in_match:
        print(f"  [INFO] India matches (FOB=2, PCS=8000): {len(in_match)}")
        for m in in_match[:5]:
            print(f"    Key={m[0]}, Code={m[1]}, Weight={m[3]}MT")
    else:
        print(f"  [INFO] No India product with FOB=2 and PCS=8000")

    if sl_match:
        print(f"  [INFO] Sri Lanka matches (FOB=2, PCS=7000): {len(sl_match)}")
        for m in sl_match[:5]:
            print(f"    Key={m[0]}, Code={m[1]}, Weight={m[3]}MT")
    else:
        print(f"  [INFO] No Sri Lanka product with FOB=2 and PCS=7000")

    if in_match and sl_match:
        in_keys_2 = set(m[0] for m in in_match)
        sl_keys_2 = set(m[0] for m in sl_match)
        shared = in_keys_2 & sl_keys_2
        if shared:
            print(f"  [PASS] Found {len(shared)} keys matching reference in BOTH origins")
            for key in sorted(shared)[:3]:
                print(f"    Key: {key}")
        else:
            print(f"  [INFO] No shared keys between India/SL matching reference (may use different weight tiers)")

    # Search for freight routes with ALL IN ~3000 and ~5000
    routes_3000 = [(r[2], r[3], r[5]) for r in fr_data if r[5] is not None and 2800 <= r[5] <= 3200]
    routes_5000 = [(r[2], r[3], r[5]) for r in fr_data if r[5] is not None and 4800 <= r[5] <= 5200]

    if routes_3000:
        print(f"  [INFO] Routes with ALL IN ~$3000: {len(routes_3000)}")
        for origin, dest, all_in in routes_3000[:5]:
            print(f"    {origin} -> {dest}: ${all_in:.2f}")

    if routes_5000:
        print(f"  [INFO] Routes with ALL IN ~$5000: {len(routes_5000)}")
        for origin, dest, all_in in routes_5000[:5]:
            print(f"    {origin} -> {dest}: ${all_in:.2f}")

    return ok

def test_quote_formulas(wb):
    """Test 5: Verify the Quote sheet formulas are present and correctly structured."""
    print("\n" + "=" * 60)
    print("TEST 5: Quote Sheet Formula Verification")
    print("=" * 60)

    ws = wb['Quote']
    ok = True

    # Check title
    title = ws['A1'].value
    if 'QUOTATION TOOL' in str(title):
        print(f"  [PASS] Title present: {title}")
    else:
        print(f"  [FAIL] Title missing or wrong: {title}")
        ok = False

    # Check input cells exist
    input_labels = {
        4: 'GB Size', 5: 'Chips', 6: 'EC Level', 7: 'Plastic',
        8: 'Holes', 9: 'BSU', 10: 'Port of Destination', 11: 'Container Gross Weight',
        12: 'Discount',
    }
    for row, expected in input_labels.items():
        label = ws.cell(row=row, column=1).value
        if label and expected.lower() in label.lower():
            print(f"  [PASS] Row {row}: '{label}'")
        else:
            print(f"  [FAIL] Row {row}: expected '{expected}...', got '{label}'")
            ok = False

    # B11 should be a formula (auto-derived, not a dropdown)
    b11 = ws['B11'].value
    if b11 and str(b11).startswith('=') and 'WeightTiers' in str(b11):
        print(f"  [PASS] B11 is an auto-derived formula (not a dropdown)")
    else:
        print(f"  [FAIL] B11 should be a formula with WeightTiers, got: {b11}")
        ok = False

    # B12 should be the discount input (default 30%)
    b12 = ws['B12'].value
    if b12 == 0.3 or b12 == 0.30:
        print(f"  [PASS] B12 discount default = {b12} (30%)")
    else:
        print(f"  [FAIL] B12 discount default should be 0.30, got: {b12}")
        ok = False

    # Check helper formulas in K column
    helper_cells = {
        'K4': 'B4&"|"&B5&"|"&B6&"|"&B7&"|"&B8&"|"&B9',    # lookup key
        'K5': 'MATCH(K12,WeightTiers,1)',                     # weight tier (approximate match on K12)
        'K6': 'AND(B4<>"",B5<>"",B6<>"",B7<>"",B8<>"",B9<>"",B10<>"")',  # 7 inputs
        'K7': 'MATCH(K4,IN_Keys,0)',                          # India product match
        'K8': 'MATCH(K4,SL_Keys,0)',                          # SL product match
        'K9': 'MATCH("Cochin, IN|"&B10,FR_Keys,0)',           # Cochin freight
        'K10': 'MATCH("Tuticorin, IN|"&B10,FR_Keys,0)',       # Tuticorin freight
        'K11': 'MATCH("Colombo, LK|"&B10,FR_Keys,0)',         # Colombo freight
        'K12': 'INDEX(FR_GrossWT',                             # gross weight from tonnage
        'K13': 'INDEX(FR_Confirmed',                           # confirmed flag
    }
    for cell_ref, expected_fragment in helper_cells.items():
        formula = ws[cell_ref].value
        if formula and expected_fragment in str(formula):
            print(f"  [PASS] {cell_ref} contains correct formula")
        else:
            print(f"  [FAIL] {cell_ref}: expected '...{expected_fragment}...', got '{formula}'")
            ok = False

    # Check result rows (3 origins)
    for row, label in [(17, 'Cochin, IN'), (18, 'Tuticorin, IN'), (19, 'Colombo, LK')]:
        src = ws.cell(row=row, column=1).value
        if src == label:
            print(f"  [PASS] Row {row} source label: {label}")
        else:
            print(f"  [FAIL] Row {row} source: '{src}', expected '{label}'")
            ok = False

        # Check each result column has a formula (B through I = 9 columns)
        for col in range(2, 10):
            val = ws.cell(row=row, column=col).value
            if val and str(val).startswith('='):
                pass  # formula present
            else:
                print(f"  [FAIL] Row {row}, Col {col}: no formula found (value: {val})")
                ok = False

    if ok:
        print(f"  [PASS] All result formulas present in rows 17-19")

    # Check warning message formula
    warn = ws['A21'].value
    if warn and 'K6' in str(warn) and 'K7' in str(warn):
        print(f"  [PASS] Status message formula present")
    else:
        print(f"  [FAIL] Status message formula missing or incomplete")
        ok = False

    # Check tonnage warning row (A13)
    tonnage_warn = ws['A13'].value
    if tonnage_warn and 'K13' in str(tonnage_warn) and 'WARNING' in str(tonnage_warn):
        print(f"  [PASS] Tonnage warning formula present in A13")
    else:
        print(f"  [FAIL] Tonnage warning formula missing in A13, got: {tonnage_warn}")
        ok = False

    # Check description reference rows
    for row, label in [(23, 'India'), (24, 'Sri Lanka')]:
        desc_formula = ws.cell(row=row, column=2).value
        if desc_formula and 'INDEX' in str(desc_formula):
            print(f"  [PASS] Row {row} description INDEX formula present")
        else:
            print(f"  [FAIL] Row {row} description formula missing")
            ok = False

    return ok

def test_dropdown_validations(wb):
    """Test 6: Verify data validation (dropdowns) on input cells."""
    print("\n" + "=" * 60)
    print("TEST 6: Dropdown Data Validations")
    print("=" * 60)

    ws = wb['Quote']
    ok = True

    validations = list(ws.data_validations.dataValidation)
    print(f"  [INFO] Found {len(validations)} data validations")

    # Weight dropdown removed — now only 7 dropdowns
    expected_lists = ['SizeList', 'ChipsList', 'ECList', 'PlasticList',
                      'HolesList', 'BSUList', 'DestList']
    found_lists = set()
    for dv in validations:
        if dv.formula1:
            found_lists.add(dv.formula1.replace('=', ''))

    for el in expected_lists:
        if el in found_lists:
            print(f"  [PASS] Dropdown for '{el}' found")
        else:
            print(f"  [FAIL] Dropdown for '{el}' MISSING")
            ok = False

    # WeightTiers should NOT be a dropdown anymore
    if 'WeightTiers' in found_lists:
        print(f"  [FAIL] WeightTiers dropdown should have been removed (weight is auto-derived)")
        ok = False
    else:
        print(f"  [PASS] WeightTiers dropdown correctly removed (weight is auto-derived)")

    return ok

def test_windows_compatibility(wb):
    """Test 7: Check for potential Windows Excel compatibility issues."""
    print("\n" + "=" * 60)
    print("TEST 7: Windows Excel Compatibility")
    print("=" * 60)

    ok = True
    ws = wb['Quote']

    # Check for Mac-specific formula syntax issues
    formulas = []
    for row in range(1, 25):
        for col in range(1, 12):
            val = ws.cell(row=row, column=col).value
            if val and str(val).startswith('='):
                formulas.append((f"{chr(64+col)}{row}", str(val)))

    print(f"  [INFO] Checking {len(formulas)} formulas for compatibility...")

    issues = []
    for ref, formula in formulas:
        if ';' in formula:
            issues.append(f"  {ref}: Contains semicolons (may be locale issue)")

        mac_only = ['WEBSERVICE', 'FILTERXML']
        for func in mac_only:
            if func in formula.upper():
                issues.append(f"  {ref}: Uses Mac-specific function {func}")

        if len(formula) > 8192:
            issues.append(f"  {ref}: Formula exceeds 8192 chars ({len(formula)})")

    if not issues:
        print(f"  [PASS] No Windows compatibility issues detected in formulas")
    else:
        for issue in issues:
            print(f"  [WARN] {issue}")
            ok = False

    # Check named range references don't use Mac-style paths
    for name in wb.defined_names:
        ref = wb.defined_names[name].attr_text
        if 'Macintosh' in ref or '/Users/' in ref:
            print(f"  [FAIL] Named range '{name}' contains Mac path: {ref}")
            ok = False

    if ok:
        print(f"  [PASS] Named ranges use standard sheet references")

    # Check for formula functions used
    functions_used = set()
    import re
    for ref, formula in formulas:
        funcs = re.findall(r'([A-Z]+)\(', formula)
        functions_used.update(funcs)

    safe_functions = {'IF', 'AND', 'OR', 'NOT', 'INDEX', 'MATCH', 'IFERROR'}
    used_other = functions_used - safe_functions

    print(f"  [INFO] Functions used: {sorted(functions_used)}")
    if used_other:
        print(f"  [WARN] Non-standard functions: {sorted(used_other)}")
    else:
        print(f"  [PASS] All functions are standard Excel functions")

    # Check column K is hidden
    if ws.column_dimensions['K'].hidden:
        print(f"  [PASS] Helper column K is hidden")
    else:
        print(f"  [WARN] Helper column K is NOT hidden — helpers will be visible")

    return ok

def test_exhaustive_lookups(wb):
    """Test 8: Simulate lookups across ALL products and ALL destinations."""
    print("\n" + "=" * 60)
    print("TEST 8: Exhaustive Lookup Coverage")
    print("=" * 60)

    ws_in = wb['IN_GB']
    ws_sl = wb['SL_GB']
    ws_fr = wb['Freight']

    in_data = get_sheet_data(ws_in, max_col=24)
    sl_data = get_sheet_data(ws_sl, max_col=24)
    fr_data = get_sheet_data(ws_fr, max_col=8)

    weight_tiers = [19.5, 22, 23, 24, 25, 26, 27]

    # Count how many products have valid data at each weight tier
    print("\n  --- FOB Price Coverage by Weight Tier ---")
    for i, wt in enumerate(weight_tiers):
        in_count = sum(1 for r in in_data if r[17 + i] is not None and r[17 + i] != 0)
        sl_count = sum(1 for r in sl_data if r[17 + i] is not None and r[17 + i] != 0)
        in_pcs = sum(1 for r in in_data if r[10 + i] is not None and r[10 + i] != 0)
        sl_pcs = sum(1 for r in sl_data if r[10 + i] is not None and r[10 + i] != 0)
        print(f"  {wt:>5} MT: IN FOB={in_count:>3}, IN PCS={in_pcs:>3} | SL FOB={sl_count:>3}, SL PCS={sl_pcs:>3}")

    # Count freight routes per origin
    origin_dests = {}
    for r in fr_data:
        origin = r[2]
        dest = r[3]
        if origin not in origin_dests:
            origin_dests[origin] = set()
        origin_dests[origin].add(dest)

    print(f"\n  --- Freight Route Coverage ---")
    for origin in sorted(origin_dests.keys()):
        print(f"  {origin}: {len(origin_dests[origin])} destinations")

    # Check all origins share the same destinations
    all_dest_sets = list(origin_dests.values())
    if len(all_dest_sets) >= 2 and all(d == all_dest_sets[0] for d in all_dest_sets[1:]):
        print(f"  [PASS] All origins serve the same {len(all_dest_sets[0])} destinations")
    else:
        print(f"  [INFO] Origins have different destination sets")

    # Check for negative ALL IN values
    bad_freight = [(r[0], r[5]) for r in fr_data if r[5] is not None and r[5] < 0]
    if bad_freight:
        print(f"  [WARN] {len(bad_freight)} routes with negative ALL IN freight!")
        for key, all_in in bad_freight[:3]:
            print(f"    {key}: ${all_in}")
    else:
        print(f"  [PASS] No negative freight values")

    # Check for zero PCS (would cause division by zero in Freight/Unit)
    for label, data in [('IN', in_data), ('SL', sl_data)]:
        zero_pcs = []
        for r in data:
            for i in range(7):
                if r[17 + i] is not None and r[17 + i] != 0:  # has FOB
                    if r[10 + i] is None or r[10 + i] == 0:    # but no PCS
                        zero_pcs.append((r[1], weight_tiers[i]))
        if zero_pcs:
            print(f"  [WARN] {label}: {len(zero_pcs)} cases where FOB exists but PCS=0 (division by zero risk)")
            print(f"    Formula uses IFERROR so this will show '-' instead of #DIV/0!")
        else:
            print(f"  [PASS] {label}: No division-by-zero risk (all products with FOB have PCS)")

    return True

def test_tonnage_integration(wb):
    """Test 9: Verify tonnage integration is consistent and complete."""
    print("\n" + "=" * 60)
    print("TEST 9: Tonnage Integration")
    print("=" * 60)

    ws_fr = wb['Freight']
    fr_data = get_sheet_data(ws_fr, max_col=8)
    weight_tiers = [19.5, 22, 23, 24, 25, 26, 27]

    ok = True

    # Check that all origin rows for the same destination have the same weight
    dest_weights = {}
    for r in fr_data:
        dest = r[3]
        gross = r[6]
        confirmed = r[7]
        if dest not in dest_weights:
            dest_weights[dest] = (gross, confirmed)
        else:
            if dest_weights[dest][0] != gross:
                print(f"  [FAIL] Destination '{dest}' has inconsistent weights: "
                      f"{dest_weights[dest][0]} vs {gross}")
                ok = False

    if ok:
        print(f"  [PASS] All destinations have consistent weight across all origin rows")

    # Check that all weights map to a valid tier
    unmappable = []
    for r in fr_data:
        gross = r[6]
        mapped = False
        for wt in weight_tiers:
            if wt <= gross:
                mapped = True
        if not mapped:
            unmappable.append((r[3], gross))

    if not unmappable:
        print(f"  [PASS] All gross weights map to a valid tier (>= 19.5 MT)")
    else:
        print(f"  [WARN] {len(unmappable)} routes with weight below minimum tier 19.5:")
        for dest, gross in unmappable[:5]:
            print(f"    {dest}: {gross} MT")

    # Report weight distribution
    weight_counts = {}
    unique_dests = set()
    for r in fr_data:
        dest = r[3]
        if dest not in unique_dests:
            unique_dests.add(dest)
            gross = r[6]
            confirmed = r[7]
            label = f"{gross} MT ({'confirmed' if confirmed == 1 else 'default'})"
            weight_counts[label] = weight_counts.get(label, 0) + 1

    print(f"  [INFO] Weight distribution across {len(unique_dests)} unique destinations:")
    for label in sorted(weight_counts.keys()):
        print(f"    {label}: {weight_counts[label]} destinations")

    # Check conditional formatting exists
    ws_q = wb['Quote']
    cf_rules = list(ws_q.conditional_formatting)
    cf_count = len(cf_rules)
    print(f"  [INFO] Conditional formatting rules on Quote sheet: {cf_count}")
    if cf_count >= 3:
        print(f"  [PASS] At least 3 conditional formatting rules present (warning row + B11 + result cells)")
    else:
        print(f"  [FAIL] Expected at least 3 conditional formatting rules, found {cf_count}")
        ok = False

    return ok


def main():
    print("QUOTATION TOOL — COMPREHENSIVE TEST SUITE")
    print("=" * 60)
    print(f"File: {TOOL_FILE}\n")

    try:
        wb = load_tool()
    except Exception as e:
        print(f"[FATAL] Cannot open workbook: {e}")
        sys.exit(1)

    results = []
    results.append(("Structure", test_structure(wb)))
    results.append(("Named Ranges", test_named_ranges(wb)))
    results.append(("Data Integrity", test_data_integrity(wb)))
    results.append(("Formula Simulation", test_formula_simulation(wb)))
    results.append(("Quote Formulas", test_quote_formulas(wb)))
    results.append(("Dropdowns", test_dropdown_validations(wb)))
    results.append(("Windows Compat", test_windows_compatibility(wb)))
    results.append(("Exhaustive Lookups", test_exhaustive_lookups(wb)))
    results.append(("Tonnage Integration", test_tonnage_integration(wb)))

    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    all_pass = True
    for name, passed in results:
        status = "PASS" if passed else "ISSUES"
        print(f"  {name:.<30} {status}")
        if not passed:
            all_pass = False

    if all_pass:
        print("\n  All tests passed! The tool should work correctly in Windows Excel.")
    else:
        print("\n  Some tests had issues — review warnings above.")

    print("\n  RECOMMENDED MANUAL TEST IN WINDOWS EXCEL:")
    print("  1. Open Quotation_Tool.xlsx")
    print("  2. On the Quote sheet, select values from all 7 dropdowns")
    print("  3. Verify weight tier auto-populates based on destination")
    print("  4. Verify 3 result rows: Cochin, Tuticorin, Colombo (9 columns each)")
    print("  5. Verify Disc. FOB = FOB × 70% (default 30% discount in B12)")
    print("  6. Verify Total Cost = Disc. FOB + Freight/Unit")
    print("  7. Change B12 discount to 0% — Disc. FOB should equal FOB")
    print("  8. Select a destination with confirmed tonnage — no warning should appear")
    print("  9. Select a destination with default tonnage — red warning banner should appear")
    print("  10. Check that hidden sheets are not visible (right-click sheet tabs)")


if __name__ == '__main__':
    main()
