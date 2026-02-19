#!/usr/bin/env python3
"""
Builds the Quotation Tool Excel workbook.
Reads Freight.xlsx and Price List Feb.xlsx, creates a self-contained
Quotation Tool.xlsx with dropdowns, formulas, and hidden data sheets.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import FormulaRule
import re
import glob
import os
import sys

# ── Configuration ──────────────────────────────────────────────────────────
# Use the folder where the script/exe is located
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Auto-detect source files in the same folder
def find_file(pattern, description):
    """Find a file matching a glob pattern in BASE_DIR."""
    matches = glob.glob(os.path.join(BASE_DIR, pattern))
    if len(matches) == 1:
        return matches[0]
    elif len(matches) > 1:
        print(f"  Found multiple {description} files:")
        for i, m in enumerate(matches, 1):
            print(f"    {i}. {os.path.basename(m)}")
        while True:
            choice = input(f"  Select {description} file (1-{len(matches)}): ").strip()
            if choice.isdigit() and 1 <= int(choice) <= len(matches):
                return matches[int(choice) - 1]
    else:
        print(f"  ERROR: No {description} file found matching '{pattern}' in:")
        print(f"    {BASE_DIR}")
        input("  Press Enter to exit...")
        sys.exit(1)

RATE_FILE = find_file('Valid from*.xlsx', 'Rate Sheet')
FREIGHT_FILE = find_file('Freight*.xlsx', 'Freight (Tonnage)')
PRICE_FILE = find_file('Price List*.xlsx', 'Price List')
OUTPUT_FILE = os.path.join(BASE_DIR, 'Quotation_Tool.xlsx')

print(f"  Rate Sheet:   {os.path.basename(RATE_FILE)}")
print(f"  Tonnage file: {os.path.basename(FREIGHT_FILE)}")
print(f"  Price List:   {os.path.basename(PRICE_FILE)}")
print()

WEIGHT_TIERS = [19.5, 22, 23, 24, 25, 26, 27]
PCS_START_COL = 11   # Column L in source (0-indexed)
FOB_START_COL = 18   # Column S in source (0-indexed)
DATA_START_ROW = 6   # Products start at row 6 in source

# ── Styles ─────────────────────────────────────────────────────────────────
TITLE_FONT = Font(name='Calibri', size=16, bold=True, color='1F4E79')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
SECTION_FONT = Font(name='Calibri', size=12, bold=True, color='1F4E79')
LABEL_FONT = Font(name='Calibri', size=11, bold=True)
INPUT_FONT = Font(name='Calibri', size=11)
RESULT_FONT = Font(name='Calibri', size=11)
WARNING_FONT = Font(name='Calibri', size=10, italic=True, color='CC0000')
DETAIL_FONT_B = Font(name='Calibri', size=9, bold=True, color='666666')
DETAIL_FONT = Font(name='Calibri', size=9, color='666666')

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
INPUT_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
ALT_ROW_FILL = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
SECTION_FILL = PatternFill(start_color='E8EEF4', end_color='E8EEF4', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)


# ── Parsing ────────────────────────────────────────────────────────────────
def parse_description(desc):
    """Parse a GB product description into its component attributes."""
    if not desc:
        return None
    d = str(desc)
    result = {
        'size': '', 'chips_pith': '', 'ec_level': '',
        'plastic': '', 'holes': 'N/A', 'bsu': 'N/A',
    }

    # Size: e.g. 100x18x16
    m = re.search(r'(\d+)\s*[xX]\s*(\d+)\s*[xX]\s*(\d+)', d)
    if m:
        result['size'] = f"{m.group(1)}X{m.group(2)}X{m.group(3)}"

    # Chips/Pith: "16mm", "6mm", "PLF 3070", "PRO 8020"
    m = re.search(r'(PLF\s*\d+|PRO\s*\d+|\d+mm)', d, re.IGNORECASE)
    if m:
        val = m.group(1).upper()
        val = re.sub(r'(PLF|PRO)\s*(\d+)', r'\1 \2', val)   # normalize spacing
        result['chips_pith'] = val

    # EC Level: NW, WA, EW, TR, FT
    m = re.search(r'\b(NW|WA|EW|TR|FT)\b', d)
    if m:
        result['ec_level'] = m.group(1)

    # Plastic duration: P1Y, P2Y, P3Y, P4Y, P5Y
    m = re.search(r'(P\d+Y)', d, re.IGNORECASE)
    if m:
        result['plastic'] = m.group(1).upper()

    # Holes
    if re.search(r'\bNO\s*HOLES\b', d, re.IGNORECASE):
        result['holes'] = 'NO HOLES'
    elif re.search(r'HOLES', d, re.IGNORECASE):
        result['holes'] = 'HOLES'

    # BSU (Bottom Side Up)
    if re.search(r'\bBSU\b', d, re.IGNORECASE):
        result['bsu'] = 'BSU'

    return result


def make_key(parsed):
    """Build a 6-field lookup key."""
    return (
        f"{parsed['size']}|{parsed['chips_pith']}|{parsed['ec_level']}|"
        f"{parsed['plastic']}|{parsed['holes']}|{parsed['bsu']}"
    )


# ── Data readers ───────────────────────────────────────────────────────────
def read_product_sheet(wb, sheet_name):
    """Read and parse one product sheet (e.g. IN-GB, SL-GB)."""
    ws = wb[sheet_name]
    products = []
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, 26)]  # A-Y
        prod_no = row[0]
        desc = row[1]
        weight = row[3]
        if not prod_no or not desc:
            continue

        parsed = parse_description(desc)
        if not parsed:
            continue

        pcs = [row[PCS_START_COL + i] if PCS_START_COL + i < len(row) else None for i in range(7)]
        fob = [row[FOB_START_COL + i] if FOB_START_COL + i < len(row) else None for i in range(7)]

        products.append({
            'key': make_key(parsed),
            'product_no': str(prod_no),
            'description': str(desc),
            'weight': weight,
            **parsed,
            'pcs': pcs,
            'fob': fob,
        })
    return products


def read_freight(rate_filepath):
    """Read ALL IN freight rates from the quarterly RATE SHEET + transit times.
    Returns all 3 origins separately (Cochin, Tuticorin, Colombo).
    Returns dict: {(origin, destination): {origin, country, destination, transit_days, all_in_usd}}
    """
    wb = openpyxl.load_workbook(rate_filepath, data_only=True)

    # ── Read ALL IN 40DRY/40HDRY from RATE SHEET ──
    ws_rate = wb['RATE SHEET']
    rates = {}
    for row in ws_rate.iter_rows(min_row=2, max_col=12, values_only=True):
        origin = row[0]   # Column A
        dest = row[1]     # Column B
        all_in = row[11]  # Column L (ALL IN 40DRY/40HDRY)
        if not origin or not dest or all_in is None:
            continue
        origin = str(origin).strip()
        dest = str(dest).strip()
        if origin == dest:
            continue
        rates[(origin, dest)] = float(all_in)

    # ── Read transit times from Transit Time sheet ──
    ws_tt = wb['Transit Time']
    transit = {}
    for row in ws_tt.iter_rows(min_row=2, values_only=True):
        receipt = row[0]   # Column A: Receipt
        delivery = row[3]  # Column D: Delivery
        tt = row[10]       # Column K: Transit Time (e.g. "46 Days")
        if not receipt or not delivery or not tt:
            continue
        key = (str(receipt).strip(), str(delivery).strip())
        if key not in transit:
            tm = re.search(r'(\d+)', str(tt))
            if tm:
                transit[key] = int(tm.group(1))

    # ── Build route map with all 3 origins ──
    route_map = {}
    for (origin, dest), all_in in rates.items():
        if ', IN' in origin:
            country = 'India'
        elif ', LK' in origin:
            country = 'Sri Lanka'
        else:
            country = origin

        transit_days = transit.get((origin, dest), 0)

        route_map[(origin, dest)] = {
            'origin': origin,
            'country': country,
            'destination': dest,
            'transit_days': transit_days,
            'all_in_usd': round(all_in, 2),
        }
    return route_map


def read_tonnage(filepath, freight_destinations):
    """Read tonnage data from the Tonnage sheet in Freight.xlsx.
    Uses only client-provided data (40HC entries).
    For duplicate ports (e.g. Japan 2-axle vs 3-axle), keeps highest gross weight.
    Returns:
        tonnage: dict {freight_destination: (gross_weight_mt, is_confirmed)}
        report: dict with quality_issues, matched, defaults lists
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Tonnage']

    # ── Step 1: Read all 40HC entries, deduplicate ────────────────────────
    raw = {}  # cleaned_port_name -> (gross_weight, original_name, row_idx)
    quality_issues = []

    for row_idx in range(3, ws.max_row + 1):
        port = ws.cell(row=row_idx, column=1).value
        country = ws.cell(row=row_idx, column=2).value
        gross = ws.cell(row=row_idx, column=5).value
        ctype = str(ws.cell(row=row_idx, column=6).value or '').strip()

        if not port or ctype != '40HC':
            continue

        port_str = str(port).strip()
        country_str = str(country).strip() if country else ''

        # Detect swapped port/country (Germany/Hamburg)
        if port_str.lower() == 'germany' and country_str.lower() == 'hamburg':
            quality_issues.append(
                f"Row {row_idx}: Port/Country swapped: '{port_str}'/'{country_str}' -> Hamburg/Germany")
            port_str = 'Hamburg'

        # Detect trailing whitespace
        if str(port).rstrip() != str(port):
            quality_issues.append(f"Row {row_idx}: Trailing whitespace in port name: '{port}'")

        # Detect misspellings (for report only — matching uses override table)
        known_misspellings = {
            'yokohoma': 'Yokohama', 'lisbao': 'Lisboa/Lisbon', 'le harve': 'Le Havre',
        }
        for wrong, correct in known_misspellings.items():
            if port_str.lower() == wrong:
                quality_issues.append(
                    f"Row {row_idx}: Possible misspelling: '{port_str}' (should be {correct})")

        # Detect annotation in name
        if '**' in port_str or '(' in port_str:
            quality_issues.append(f"Row {row_idx}: Annotation in port name: '{port_str}'")

        if gross is None:
            quality_issues.append(f"Row {row_idx}: Missing gross weight for '{port_str}'")
            continue

        gross_val = float(gross)
        key = port_str.lower()

        if key in raw:
            if gross_val > raw[key][0]:
                quality_issues.append(
                    f"Duplicate port '{port_str}' (row {row_idx}): keeping higher weight "
                    f"{gross_val} over {raw[key][0]}")
                raw[key] = (gross_val, port_str, row_idx)
            else:
                quality_issues.append(
                    f"Duplicate port '{port_str}' (row {row_idx}): skipping lower weight "
                    f"{gross_val} (keeping {raw[key][0]})")
        else:
            raw[key] = (gross_val, port_str, row_idx)

    # ── Step 2: Match tonnage ports to freight destinations ───────────────
    # Override table for known name mismatches (freight_dest_lower -> tonnage_key)
    OVERRIDES = {
        'london gateway, gb': 'london gateway terminal',
        'las palmas de gran canaria, es': 'las palmas',
        'lisbon, pt': 'lisbao',
        'guayaquil-posorja, ec': 'guayaquil',
        'cartagena, es': 'cartagena',
    }

    # Cartagena, CO: tonnage says Spain, freight says Colombia — do NOT match
    SKIP_MATCHES = {'cartagena, co'}

    tonnage = {}
    matched = []
    defaults = []

    for dest in freight_destinations:
        dest_lower = dest.lower()

        if dest_lower in SKIP_MATCHES:
            quality_issues.append(
                f"Cartagena: tonnage lists Spain but freight destination is '{dest}' (Colombia) — not matched")
            tonnage[dest] = (23, False)
            defaults.append(dest)
            continue

        # Try override first
        tonnage_key = OVERRIDES.get(dest_lower)

        # Try extracting city name from "City, CC" format
        if tonnage_key is None:
            city = dest.split(',')[0].strip().lower()
            if city in raw:
                tonnage_key = city

        if tonnage_key and tonnage_key in raw:
            gross_val = raw[tonnage_key][0]
            tonnage[dest] = (gross_val, True)
            matched.append((dest, raw[tonnage_key][1], gross_val))
        else:
            tonnage[dest] = (23, False)
            defaults.append(dest)

    report = {
        'quality_issues': quality_issues,
        'matched': matched,
        'defaults': defaults,
        'matched_count': len(matched),
        'default_count': len(defaults),
        'total': len(freight_destinations),
    }
    return tonnage, report


# ── Workbook builder ───────────────────────────────────────────────────────
def build_tool(in_products, sl_products, freight, tonnage):
    wb = openpyxl.Workbook()

    # ── 1. Collect unique values for dropdowns ─────────────────────────────
    all_products = in_products + sl_products
    sizes = sorted({p['size'] for p in all_products if p['size']})
    chips = sorted({p['chips_pith'] for p in all_products if p['chips_pith']})
    ecs = sorted({p['ec_level'] for p in all_products if p['ec_level']})
    plastics = sorted({p['plastic'] for p in all_products if p['plastic']})
    holes_vals = sorted({p['holes'] for p in all_products})      # HOLES, NO HOLES, N/A
    bsu_vals = sorted({p['bsu'] for p in all_products})           # BSU, N/A
    destinations = sorted({v['destination'] for v in freight.values()})

    # ── 2. Create Lists sheet ──────────────────────────────────────────────
    ws_lists = wb.active
    ws_lists.title = 'Lists'
    list_cols = {
        'A': ('Size', sizes),
        'B': ('Chips_Pith', chips),
        'C': ('EC_Level', ecs),
        'D': ('Plastic', plastics),
        'E': ('Holes', holes_vals),
        'F': ('BSU', bsu_vals),
        'G': ('Destination', destinations),
        'H': ('Weight_MT', WEIGHT_TIERS),
    }
    for col_letter, (header, values) in list_cols.items():
        col_idx = ord(col_letter) - ord('A') + 1
        ws_lists.cell(row=1, column=col_idx, value=header)
        for r, val in enumerate(values, 2):
            ws_lists.cell(row=r, column=col_idx, value=val)

    # ── 3. Write product data sheets ───────────────────────────────────────
    ws_in = wb.create_sheet('IN_GB')
    _write_product_sheet(ws_in, in_products)

    ws_sl = wb.create_sheet('SL_GB')
    _write_product_sheet(ws_sl, sl_products)

    # ── 4. Write freight sheet ─────────────────────────────────────────────
    ws_fr = wb.create_sheet('Freight')
    fr_headers = ['Key', 'Country', 'Origin', 'Destination', 'Transit_Days',
                  'All_In_USD', 'Gross_Weight_MT', 'Weight_Confirmed']
    for c, h in enumerate(fr_headers, 1):
        ws_fr.cell(row=1, column=c, value=h)
    fr_rows = sorted(freight.values(), key=lambda x: (x['origin'], x['destination']))
    for r, fr in enumerate(fr_rows, 2):
        dest = fr['destination']
        gross_wt, confirmed = tonnage.get(dest, (23, False))
        ws_fr.cell(row=r, column=1, value=f"{fr['origin']}|{fr['destination']}")
        ws_fr.cell(row=r, column=2, value=fr['country'])
        ws_fr.cell(row=r, column=3, value=fr['origin'])
        ws_fr.cell(row=r, column=4, value=fr['destination'])
        ws_fr.cell(row=r, column=5, value=fr['transit_days'])
        ws_fr.cell(row=r, column=6, value=fr['all_in_usd'])
        ws_fr.cell(row=r, column=7, value=gross_wt)
        ws_fr.cell(row=r, column=8, value=1 if confirmed else 0)
    fr_last = len(fr_rows) + 1

    # ── 5. Define named ranges ─────────────────────────────────────────────
    in_last = len(in_products) + 1
    sl_last = len(sl_products) + 1

    # Data sheet columns: A=Key B=ProdNo C=Desc D=Weight E=Size F=Chips
    #   G=EC H=Plastic I=Holes J=BSU  K-Q=PCS  R-X=FOB
    named = {
        'IN_Keys':    f"IN_GB!$A$2:$A${in_last}",
        'IN_ProdNos': f"IN_GB!$B$2:$B${in_last}",
        'IN_Descs':   f"IN_GB!$C$2:$C${in_last}",
        'IN_PCS':     f"IN_GB!$K$2:$Q${in_last}",
        'IN_FOB':     f"IN_GB!$R$2:$X${in_last}",
        'SL_Keys':    f"SL_GB!$A$2:$A${sl_last}",
        'SL_ProdNos': f"SL_GB!$B$2:$B${sl_last}",
        'SL_Descs':   f"SL_GB!$C$2:$C${sl_last}",
        'SL_PCS':     f"SL_GB!$K$2:$Q${sl_last}",
        'SL_FOB':     f"SL_GB!$R$2:$X${sl_last}",
        'FR_Keys':      f"Freight!$A$2:$A${fr_last}",
        'FR_Transit':   f"Freight!$E$2:$E${fr_last}",
        'FR_AllIn':     f"Freight!$F$2:$F${fr_last}",
        'FR_GrossWT':   f"Freight!$G$2:$G${fr_last}",
        'FR_Confirmed': f"Freight!$H$2:$H${fr_last}",
        'SizeList':    f"Lists!$A$2:$A${len(sizes)+1}",
        'ChipsList':   f"Lists!$B$2:$B${len(chips)+1}",
        'ECList':      f"Lists!$C$2:$C${len(ecs)+1}",
        'PlasticList': f"Lists!$D$2:$D${len(plastics)+1}",
        'HolesList':   f"Lists!$E$2:$E${len(holes_vals)+1}",
        'BSUList':     f"Lists!$F$2:$F${len(bsu_vals)+1}",
        'DestList':    f"Lists!$G$2:$G${len(destinations)+1}",
        'WeightTiers': f"Lists!$H$2:$H${len(WEIGHT_TIERS)+1}",
    }
    for name, ref in named.items():
        dn = DefinedName(name, attr_text=ref)
        wb.defined_names.add(dn)

    # ── 6. Create Quote sheet ──────────────────────────────────────────────
    ws_q = wb.create_sheet('Quote')
    wb.move_sheet('Quote', offset=-4)
    _build_quote_sheet(ws_q)

    # ── 7. Hide data sheets ────────────────────────────────────────────────
    ws_lists.sheet_state = 'hidden'
    ws_in.sheet_state = 'hidden'
    ws_sl.sheet_state = 'hidden'
    ws_fr.sheet_state = 'hidden'

    return wb


def _write_product_sheet(ws, products):
    """Write parsed product data to a hidden sheet.
    Columns: A=Key B=ProdNo C=Desc D=Weight E=Size F=Chips G=EC H=Plastic
             I=Holes J=BSU  K-Q=PCS(7 tiers)  R-X=FOB(7 tiers)
    """
    headers = [
        'Key', 'Product_No', 'Description', 'Weight',
        'Size', 'Chips_Pith', 'EC_Level', 'Plastic', 'Holes', 'BSU',
        'PCS@19.5', 'PCS@22', 'PCS@23', 'PCS@24', 'PCS@25', 'PCS@26', 'PCS@27',
        'FOB@19.5', 'FOB@22', 'FOB@23', 'FOB@24', 'FOB@25', 'FOB@26', 'FOB@27',
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    for r, p in enumerate(products, 2):
        ws.cell(row=r, column=1, value=p['key'])
        ws.cell(row=r, column=2, value=p['product_no'])
        ws.cell(row=r, column=3, value=p['description'])
        ws.cell(row=r, column=4, value=p['weight'])
        ws.cell(row=r, column=5, value=p['size'])
        ws.cell(row=r, column=6, value=p['chips_pith'])
        ws.cell(row=r, column=7, value=p['ec_level'])
        ws.cell(row=r, column=8, value=p['plastic'])
        ws.cell(row=r, column=9, value=p['holes'])
        ws.cell(row=r, column=10, value=p['bsu'])
        for i in range(7):
            ws.cell(row=r, column=11 + i, value=p['pcs'][i])   # K-Q
            ws.cell(row=r, column=18 + i, value=p['fob'][i])   # R-X


def _build_quote_sheet(ws):
    """Build the user-facing Quote sheet with inputs, outputs, and formulas."""

    # Column widths
    col_widths = {'A': 30, 'B': 24, 'C': 20, 'D': 20, 'E': 20,
                  'F': 18, 'G': 18, 'H': 20, 'I': 20, 'J': 5, 'K': 22}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── Title ──────────────────────────────────────────────────────────────
    ws.merge_cells('A1:I1')
    ws['A1'].value = 'QUOTATION TOOL  --  GB Products'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 35

    # ── Input section ──────────────────────────────────────────────────────
    ws.merge_cells('A3:B3')
    ws['A3'].value = 'INPUT PARAMETERS'
    ws['A3'].font = SECTION_FONT
    ws['A3'].fill = SECTION_FILL
    ws['B3'].fill = SECTION_FILL

    # Dropdown inputs  (row, label, named_range)
    inputs = [
        (4,  'GB Size',                     'SizeList'),
        (5,  'Chips / Pith Ratio',          'ChipsList'),
        (6,  'EC Level',                    'ECList'),
        (7,  'Plastic Duration',            'PlasticList'),
        (8,  'Holes',                       'HolesList'),
        (9,  'BSU (Bottom Side Up)',        'BSUList'),
        (10, 'Port of Destination',         'DestList'),
    ]
    for row, label, list_name in inputs:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        cell_b = ws.cell(row=row, column=2)
        cell_b.fill = INPUT_FILL
        cell_b.font = INPUT_FONT
        cell_b.border = THIN_BORDER
        cell_b.alignment = Alignment(horizontal='center')
        dv = DataValidation(type='list', formula1=f'={list_name}', allow_blank=True)
        dv.error = 'Please select a value from the dropdown list.'
        dv.errorTitle = 'Invalid Input'
        dv.prompt = f'Select {label}'
        dv.promptTitle = label
        ws.add_data_validation(dv)
        dv.add(cell_b)

    # Row 11: Auto-derived weight tier (display-only, not a dropdown)
    DISPLAY_FILL = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    ws.cell(row=11, column=1, value='Container Gross Weight (MT)').font = LABEL_FONT
    wt_cell = ws.cell(row=11, column=2)
    wt_cell.value = '=IF(B10="","",IF(K5=0,"N/A",INDEX(WeightTiers,K5)&" MT"&IF(K13=0," (UNCONFIRMED)","")))'
    wt_cell.fill = DISPLAY_FILL
    wt_cell.font = Font(name='Calibri', size=11, italic=True)
    wt_cell.border = THIN_BORDER
    wt_cell.alignment = Alignment(horizontal='center')

    # Row 12: Discount percentage (manual entry, default 30%)
    ws.cell(row=12, column=1, value='Discount (%)').font = LABEL_FONT
    disc = ws.cell(row=12, column=2, value=0.30)
    disc.fill = INPUT_FILL
    disc.font = INPUT_FONT
    disc.border = THIN_BORDER
    disc.alignment = Alignment(horizontal='center')
    disc.number_format = '0%'

    # ── Helper cells (column K, hidden) ────────────────────────────────────
    # K4: lookup key (6 fields)
    ws['K4'] = '=B4&"|"&B5&"|"&B6&"|"&B7&"|"&B8&"|"&B9'
    # K5: weight tier index — approximate match (largest tier <= gross weight)
    ws['K5'] = '=IFERROR(MATCH(K12,WeightTiers,1),0)'
    # K6: all required inputs filled? (7 inputs — weight is auto-derived)
    ws['K6'] = '=AND(B4<>"",B5<>"",B6<>"",B7<>"",B8<>"",B9<>"",B10<>"")'
    # K7: India product match row
    ws['K7'] = '=IFERROR(MATCH(K4,IN_Keys,0),0)'
    # K8: Sri Lanka product match row
    ws['K8'] = '=IFERROR(MATCH(K4,SL_Keys,0),0)'
    # K9: Cochin freight match row
    ws['K9'] = '=IFERROR(MATCH("Cochin, IN|"&B10,FR_Keys,0),0)'
    # K10: Tuticorin freight match row
    ws['K10'] = '=IFERROR(MATCH("Tuticorin, IN|"&B10,FR_Keys,0),0)'
    # K11: Colombo freight match row
    ws['K11'] = '=IFERROR(MATCH("Colombo, LK|"&B10,FR_Keys,0),0)'
    # K12: raw gross weight from tonnage (cascade: Cochin -> Tuticorin -> Colombo)
    ws['K12'] = '=IF(K9>0,INDEX(FR_GrossWT,K9),IF(K10>0,INDEX(FR_GrossWT,K10),IF(K11>0,INDEX(FR_GrossWT,K11),0)))'
    # K13: weight confirmed flag (1=client data, 0=default)
    ws['K13'] = '=IF(K9>0,INDEX(FR_Confirmed,K9),IF(K10>0,INDEX(FR_Confirmed,K10),IF(K11>0,INDEX(FR_Confirmed,K11),0)))'
    ws.column_dimensions['K'].hidden = True

    # ── Tonnage warning row ───────────────────────────────────────────────
    ws.merge_cells('A13:I13')
    ws['A13'] = (
        '=IF(AND(B10<>"",K13=0),'
        '"WARNING: No confirmed weight data for this destination. '
        'Using default 23 MT. Verify before quoting.","")'
    )
    ws['A13'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    # Conditional formatting: orange background when warning is active
    ws.conditional_formatting.add('A13:I13', FormulaRule(
        formula=['AND($B$10<>"",$K$13=0)'],
        fill=PatternFill(start_color='E67E22', end_color='E67E22', fill_type='solid'),
        font=Font(bold=True, color='FFFFFF'),
    ))
    # Conditional formatting: light orange on B11 when unconfirmed
    ws.conditional_formatting.add('B11', FormulaRule(
        formula=['AND($B$10<>"",$K$13=0)'],
        fill=PatternFill(start_color='FDEBD0', end_color='FDEBD0', fill_type='solid'),
        font=Font(bold=True, color='E67E22'),
    ))

    # ── Results section ────────────────────────────────────────────────────
    ws.row_dimensions[14].height = 8   # spacer
    ws.merge_cells('A15:I15')
    ws['A15'].value = 'RESULTS'
    ws['A15'].font = SECTION_FONT
    ws['A15'].fill = SECTION_FILL
    for c in range(2, 10):
        ws.cell(row=15, column=c).fill = SECTION_FILL

    # Column headers
    result_headers = [
        'Source', 'Product Code', 'FOB Price / Unit ($)',
        'Disc. FOB / Unit ($)', 'Freight / Container ($)', 'Units / Container',
        'Freight / Unit ($)', 'Total Cost / Unit ($)', 'Transit Time (Days)',
    ]
    for c, h in enumerate(result_headers, 1):
        cell = ws.cell(row=16, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[16].height = 32

    # Cochin row (17)
    _write_result_row(ws, 17, 'Cochin, IN', 'IN', 'K9', is_alt=False)
    # Tuticorin row (18)
    _write_result_row(ws, 18, 'Tuticorin, IN', 'IN', 'K10', is_alt=True)
    # Colombo row (19)
    _write_result_row(ws, 19, 'Colombo, LK', 'SL', 'K11', is_alt=False)

    # Conditional formatting: amber tint on weight-dependent result cells when unconfirmed
    amber_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    for col_range in ['C17:C19', 'D17:D19', 'F17:F19', 'H17:H19']:
        ws.conditional_formatting.add(col_range, FormulaRule(
            formula=['AND($B$10<>"",$K$13=0)'],
            fill=amber_fill,
        ))

    # ── Status message ─────────────────────────────────────────────────────
    ws.row_dimensions[20].height = 8
    ws.merge_cells('A21:I21')
    ws['A21'] = (
        '=IF(NOT(K6),"Please fill in all input fields above.",'
        'IF(AND(K7=0,K8=0),"No results found with the query inputs.",'
        'IF(AND(K7>0,K8>0,K9>0,K10>0,K11>0),"",'
        'IF(AND(K7=0,K8>0),"Product not available from India.",'
        'IF(AND(K7>0,K8=0),"Product not available from Sri Lanka.",'
        'IF(AND(K9=0,K10=0,K11>0),"No freight routes available from India to this destination.",'
        'IF(AND(K9>0,K10>0,K11=0),"No freight route available from Sri Lanka to this destination.",'
        'IF(AND(K9=0,K10=0,K11=0),"No freight routes available to this destination.",""))))))))'
    )
    ws['A21'].font = WARNING_FONT

    # ── Full descriptions (reference rows) ─────────────────────────────────
    ws.row_dimensions[22].height = 8
    ws['A23'].value = 'India - Full Description'
    ws['A23'].font = DETAIL_FONT_B
    ws.merge_cells('B23:I23')
    ws['B23'] = '=IF(NOT(K6),"",IF(K7=0,"-",INDEX(IN_Descs,K7)))'
    ws['B23'].font = DETAIL_FONT

    ws['A24'].value = 'Sri Lanka - Full Description'
    ws['A24'].font = DETAIL_FONT_B
    ws.merge_cells('B24:I24')
    ws['B24'] = '=IF(NOT(K6),"",IF(K8=0,"-",INDEX(SL_Descs,K8)))'
    ws['B24'].font = DETAIL_FONT

    # Page setup
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)


def _write_result_row(ws, row, label, prefix, fr, is_alt=False):
    """Write one result row (Cochin/Tuticorin/Colombo) with lookup formulas."""
    fill = ALT_ROW_FILL if is_alt else PatternFill()
    mr = 'K7' if prefix == 'IN' else 'K8'     # product match ref

    def cell(col, value, fmt=None, align='center'):
        c = ws.cell(row=row, column=col, value=value)
        c.font = RESULT_FONT
        c.fill = fill
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal=align)
        if fmt:
            c.number_format = fmt
        return c

    # A: Source label
    c = ws.cell(row=row, column=1, value=label)
    c.font = LABEL_FONT
    c.fill = fill
    c.border = THIN_BORDER

    # B: Product Code
    cell(2, f'=IF(NOT($K$6),"",IF({mr}=0,"-",INDEX({prefix}_ProdNos,{mr})))')

    # C: FOB Price / Unit
    cell(3, f'=IF(NOT($K$6),"",IF(OR({mr}=0,$K$5=0),"-",INDEX({prefix}_FOB,{mr},$K$5)))', '#,##0.00')

    # D: Discounted FOB / Unit  = C × (1 - discount%)
    cell(4, f'=IF(NOT($K$6),"",IF(OR({mr}=0,$K$5=0),"-",C{row}*(1-$B$12)))', '#,##0.00')

    # E: Freight / Container (ALL IN rate × 1.0605 insurance+margin multiplier)
    cell(5, f'=IF(NOT($K$6),"",IF({fr}=0,"-",INDEX(FR_AllIn,{fr})*1.0605))', '#,##0.00')

    # F: Units / Container
    cell(6, f'=IF(NOT($K$6),"",IF(OR({mr}=0,$K$5=0),"-",INDEX({prefix}_PCS,{mr},$K$5)))', '#,##0')

    # G: Freight / Unit  = E/F
    cell(7, f'=IF(NOT($K$6),"",IFERROR(E{row}/F{row},"-"))', '#,##0.000')

    # H: Total Cost / Unit  = D (discounted FOB) + G (freight/unit)
    cell(8, f'=IF(NOT($K$6),"",IFERROR(D{row}+G{row},"-"))', '#,##0.000')

    # I: Transit Time (Days)
    cell(9, f'=IF(NOT($K$6),"",IF({fr}=0,"-",INDEX(FR_Transit,{fr})))', '0')


# ── Main ───────────────────────────────────────────────────────────────────
def main():
    print("Reading Price List...")
    price_wb = openpyxl.load_workbook(PRICE_FILE)
    in_products = read_product_sheet(price_wb, 'IN-GB')
    sl_products = read_product_sheet(price_wb, 'SL-GB')
    print(f"  IN-GB: {len(in_products)} products")
    print(f"  SL-GB: {len(sl_products)} products")

    # Report duplicate keys
    for label, products in [('IN-GB', in_products), ('SL-GB', sl_products)]:
        keys = [p['key'] for p in products]
        dupes = {k for k in keys if keys.count(k) > 1}
        if dupes:
            print(f"  Note: {label} has {len(dupes)} duplicate keys (first match used):")
            for d in sorted(dupes)[:10]:
                print(f"    {d}")
            if len(dupes) > 10:
                print(f"    ... and {len(dupes) - 10} more")

    print("Reading Freight data from Rate Sheet...")
    freight = read_freight(RATE_FILE)
    origins = sorted({v['origin'] for v in freight.values()})
    print(f"  {len(freight)} routes across {len(origins)} origins: {', '.join(origins)}")

    print("Reading Tonnage data...")
    freight_destinations = sorted({v['destination'] for v in freight.values()})
    tonnage, tonnage_report = read_tonnage(FREIGHT_FILE, freight_destinations)

    if tonnage_report['quality_issues']:
        print("\n  === TONNAGE DATA QUALITY REPORT ===")
        for issue in tonnage_report['quality_issues']:
            print(f"  WARNING: {issue}")

    print(f"\n  Tonnage matched: {tonnage_report['matched_count']}/{tonnage_report['total']} "
          f"destinations (client-confirmed)")
    print(f"  Using default 23 MT: {tonnage_report['default_count']}/{tonnage_report['total']} "
          f"destinations (UNCONFIRMED)")
    if tonnage_report['defaults']:
        print("  Destinations using default weight:")
        for d in sorted(tonnage_report['defaults']):
            print(f"    - {d}")
    print()

    print("Building Quotation Tool workbook...")
    wb = build_tool(in_products, sl_products, freight, tonnage)

    print(f"Saving to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print()
    print("Done! Open Quotation_Tool.xlsx in Excel and use the dropdowns on the Quote sheet.")
    input("Press Enter to exit...")


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"\nERROR: {e}")
        input("Press Enter to exit...")
